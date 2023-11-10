import tkinter as tk
from tkinter import ttk
import os
import openpyxl
from datetime import datetime, timedelta

def save_data(event=None):
    date = date_entry.get()
    month = month_entry.get()
    flat_no = flat_no_entry.get()
    building_no = building_no_entry.get()
    charges = charges_entry.get()

    if date == '':
        date = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        month = date[:7]

    data = {
        "Date": [date],
        "Month": [month],
        "Flat No": [flat_no],
        "Building No": [building_no],
        "Charges": [charges]
    }

    filename = 'data_entry.xlsx'
    sheetname = month

    if not os.path.isfile(filename):
        wb = openpyxl.Workbook()
        wb.active.title = sheetname
        for idx, col_name in enumerate(data.keys(), 1):
            wb.active.cell(row=1, column=idx, value=col_name)
        wb.save(filename)
        
    wb = openpyxl.load_workbook(filename)
    if sheetname not in wb.sheetnames:
        wb.create_sheet(sheetname)
        for idx, col_name in enumerate(data.keys(), 1):
            wb[sheetname].cell(row=1, column=idx, value=col_name)
    current_max_row = wb[sheetname].max_row
    for idx, value in enumerate(data.values(), 1):
        wb[sheetname].cell(row=current_max_row+1, column=idx, value=value[0])
    wb.save(filename)
    wb.close()

    status_label.config(text="Data saved successfully")

    # Clear entry fields
    flat_no_entry.delete(0, 'end')
    building_no_entry.delete(0, 'end')
    charges_entry.delete(0, 'end')

# Create the main window
root = tk.Tk()
root.title("Data Entry Form")

# Create and configure the frame
frame = ttk.Frame(root)
frame.grid(row=0, column=0, padx=20, pady=20)

# Labels and Entry fields
date_label = ttk.Label(frame, text="Date")
date_label.grid(row=0, column=0)
date_entry = ttk.Entry(frame)
date_entry.grid(row=0, column=1)

# Set the date to the next day by default
default_date = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
date_entry.insert(0, default_date)

month_label = ttk.Label(frame, text="Month")
month_label.grid(row=1, column=0)
month_entry = ttk.Entry(frame)
month_entry.grid(row=1, column=1)

# Set the month to be the same as the date in English
default_month = (datetime.now() + timedelta(days=1)).strftime("%B")
month_entry.insert(0, default_month)

flat_no_label = ttk.Label(frame, text="Flat No")
flat_no_label.grid(row=2, column=0)
flat_no_entry = ttk.Entry(frame)
flat_no_entry.grid(row=2, column=1)

building_no_label = ttk.Label(frame, text="Building No")
building_no_label.grid(row=3, column=0)
building_no_entry = ttk.Entry(frame)
building_no_entry.grid(row=3, column=1)

charges_label = ttk.Label(frame, text="Charges")
charges_label.grid(row=4, column=0)
charges_entry = ttk.Entry(frame)
charges_entry.grid(row=4, column=1)

# Bind the Enter key to the save_data function
root.bind('<Return>', save_data)

# Save button
save_button = ttk.Button(frame, text="Save", command=save_data)
save_button.grid(row=5, column=0, columnspan=2)

# Status label
status_label = ttk.Label(frame, text="")
status_label.grid(row=6, column=0, columnspan=2)

# Start the GUI application
root.mainloop()
